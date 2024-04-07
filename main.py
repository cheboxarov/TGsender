from time import sleep
import socks
import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton
import telethon
from telethon import TelegramClient, sync
import asyncio
import os
import socks
import openpyxl

bot = telebot.TeleBot('6021346982:AAFaS3Y_Hlhwzts0-AuUX1vmjJeXrqX4l9M')

user_data = {}
stop_spam_process = False

def main():
    bot.polling()


def create_client_dir(chat_id):
    if not os.path.isdir("clients"):
        os.mkdir("clients")
    if not os.path.isdir("clients/" + str(chat_id)):
        os.mkdir("clients/" + str(chat_id))
    if not os.path.isdir("clients/" + str(chat_id) + "/users"):
        os.mkdir("clients/" + str(chat_id) + "/users")
    if not os.path.isdir("clients/" + str(chat_id) + "/sessions"):
        os.mkdir("clients/" + str(chat_id) + "/sessions")


@bot.message_handler(commands=['start'])
def start(message):
    chat_id = message.chat.id
    create_client_dir(chat_id)
    bot.send_message(message.chat.id, 'Привет! Этот бот предназначен для рассылки сообщений. '
                                      'Используйте кнопки ниже для управления.', reply_markup=create_keyboard())


@bot.message_handler(commands=['user_append'])
def append_user(message):
    bot.send_message(message.chat.id,
                     "Заполните значения через пробел:\nAPI_ID API_HASH LOGIN PASSWORD\nAPI_ID AI_HASH LOGIN PASSWORD SOCKS5\nИли напишите 'отмена'")
    bot.register_next_step_handler(message, append_user_handler)


def append_user_handler(message):
    chat_id = message.chat.id
    appended_user = message.text.split(" ")
    if len(appended_user) < 4:
        bot.send_message(chat_id, "API_ID API_HASH LOGIN PASSWORD")
        if message.text != "отмена":
            bot.register_next_step_handler(message, append_user_handler)
        return
    create_client_dir(chat_id)
    file_path = "clients/" + str(chat_id) + "/users/" + appended_user[2] + ".clnt"
    file = open(file_path, "w")
    user_str = "api_id=" + appended_user[0]
    user_str += "\napi_hash=" + appended_user[1]
    user_str += "\nlogin=" + appended_user[2]
    user_str += "\npassword=" + appended_user[3]
    if len(appended_user) == 5:
        user_str += "\nproxy=" + appended_user[4]
    file.write(user_str)
    user = {}
    user["api_id"] = appended_user[0]
    user["api_hash"] = appended_user[1]
    user["login"] = appended_user[2]
    user["password"] = appended_user[3]
    if len(appended_user) == 5:
        user["proxy"] = appended_user[4]
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            login = user["proxy"].split("@")[0].split(":")[0]
            password = user["proxy"].split("@")[0].split(":")[1]
            ip = user["proxy"].split("@")[1].split(":")[0]
            port = user["proxy"].split("@")[1].split(":")[1]
            proxy = {
                'proxy_type': 'socks5',  # (mandatory) protocol to use (see above)
                'addr': ip,  # (mandatory) proxy IP address
                'port': int(port),  # (mandatory) proxy port number
                'username': login,  # (optional) username if the proxy requires auth
                'password': password,  # (optional) password if the proxy requires auth
                'rdns': True  # (optional) whether to use remote or local resolve, default remote
            }
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"],
                                       proxy=proxy)
        except KeyError:
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"])
        tg_client.connect()
        if tg_client.is_connected():
            tg_client.send_code_request(user["login"], force_sms=False)
            bot.send_message(chat_id, "Введите код.")
            bot.register_next_step_handler(message, code_handler, loop, user, tg_client)
    except BaseException as err:
        bot.send_message(chat_id, "Аккаунт " + user["login"] + " не доступен: " + str(err))


def code_handler(message, loop, user, tg_client):
    chat_id = message.chat.id
    asyncio.set_event_loop(loop)
    try:
        try:
            tg_client.sign_in(user["login"], code=message.text)
        except telethon.errors.SessionPasswordNeededError:
            tg_client.sign_in(password=user["password"])
        bot.send_message(chat_id, "Добавлен пользователь " + user["login"])
    except BaseException as err:
        bot.send_message(chat_id, "Ошибка авторизации: " + str(err))
    tg_client.disconnect()


@bot.message_handler(commands=['get_users'])
def get_users(message):
    chat_id = message.chat.id
    users = get_saved_users(message.chat.id)
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    for user in users:
        text = "Номер телефона: " + user["login"]
        text += "\nПароль: " + user["password"]
        text += "\nAPI_ID: " + user["api_id"]
        text += "\nAPI_HASH: " + user["api_hash"]
        try:
            text += "\nProxy: " +user["proxy"]
        except KeyError:
            pass
        if check_user(user, chat_id):
            text += "\nАккаунт валидный"
        else:
            text += "\nАккаунт не валидный"
        bot.send_message(message.chat.id, text)

def check_user(user, chat_id):
    try:
        try:
            login = user["proxy"].split("@")[0].split(":")[0]
            password = user["proxy"].split("@")[0].split(":")[1]
            ip = user["proxy"].split("@")[1].split(":")[0]
            port = user["proxy"].split("@")[1].split(":")[1]
            proxy = {
                'proxy_type': 'socks5',  # (mandatory) protocol to use (see above)
                'addr': ip,  # (mandatory) proxy IP address
                'port': int(port),  # (mandatory) proxy port number
                'username': login,  # (optional) username if the proxy requires auth
                'password': password,  # (optional) password if the proxy requires auth
                'rdns': True  # (optional) whether to use remote or local resolve, default remote
            }
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"],
                                       proxy=proxy)
        except KeyError:
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"])
        tg_client.start(user["login"], password=user["password"])
        tg_client.disconnect()
        return True
    except BaseException as err:
        return False

@bot.message_handler(commands=['set_message'])
def set_message(message):
    bot.send_message(message.chat.id, 'Введите сообщение для рассылки:')
    bot.register_next_step_handler(message, set_message_handler)


def set_message_handler(message):
    chat_id = message.chat.id
    message_text = message.text
    user_data['message'] = message_text
    bot.send_message(chat_id, f'Сообщение для рассылки успешно установлено: {message_text}')


@bot.message_handler(commands=['set_recipients'])
def set_recipients(message):
    bot.send_message(message.chat.id, 'Введите список получателей через пробел или запятую:')
    bot.register_next_step_handler(message, set_recipients_handler)


def set_recipients_handler(message):
    chat_id = message.chat.id
    recipients = message.text.split()
    user_data['recipients'] = recipients
    bot.send_message(chat_id, f'Список получателей успешно установлен: {recipients}')

@bot.message_handler(commands=["stop_spam"])
def stop_spam(message):
    global stop_spam_process
    stop_spam_process = True
    bot.send_message(message.chat.id, "Останавливаю")

@bot.message_handler(commands=['send_spam'])
def send_spam(message):
    chat_id = message.chat.id
    try:
        message_text = user_data["message"]
    except:
        bot.send_message(chat_id, "Укажите сообщение /set_message")
        return
    recipients = []
    try:
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("clients/" + str(chat_id) + "/base.xlsx")
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values
        for i in range(1, worksheet.max_row):
            for col in worksheet.iter_cols(3, 3):
                recipients.append(col[i].value)
    except Exception as e:
        bot.send_message(chat_id, 'База не найдена, используйте /set_base ' + str(e))
        return
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    users = get_saved_users(chat_id)
    valid_users = []
    for user in get_saved_users(chat_id):
        if check_user(user, chat_id):
            valid_users.append(user)
        else:
            bot.send_message(chat_id, "Аккаунт " + user["login"] + " - не валид")
    _it = 0
    _itRec = 0
    _errors = 0
    while _itRec < len(recipients):
        global stop_spam_process
        if stop_spam_process:
            stop_spam_process = False
            return
        user = valid_users[_it]
        _it += 1
        if _it == len(valid_users):
            _it = 0
        if recipients[_itRec] == None:
            _itRec += 1
            continue
        file = open("clients/" + str(chat_id) + "/sended.txt", "r")
        sended_text = file.read()
        if recipients[_itRec] in sended_text:
            _itRec += 1
            file.close()
            continue
        file.close()
        try:
            login = user["proxy"].split("@")[0].split(":")[0]
            password = user["proxy"].split("@")[0].split(":")[1]
            ip = user["proxy"].split("@")[1].split(":")[0]
            port = user["proxy"].split("@")[1].split(":")[1]
            proxy = {
                'proxy_type': 'socks5',  # (mandatory) protocol to use (see above)
                'addr': ip,  # (mandatory) proxy IP address
                'port': int(port),  # (mandatory) proxy port number
                'username': login,  # (optional) username if the proxy requires auth
                'password': password,  # (optional) password if the proxy requires auth
                'rdns': True  # (optional) whether to use remote or local resolve, default remote
            }
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"],
                                       proxy=proxy)
        except KeyError:
            tg_client = TelegramClient("clients/" + str(chat_id) + "/sessions/" + str(user["login"]),
                                       int(user["api_id"]), user["api_hash"])
        tg_client.start(user["login"], user["password"])
        try:
            file = open("clients/" + str(chat_id) + "/sended.txt", "r")
            sended_text = file.read()
            if not (recipients[_itRec] in sended_text):
                print(user["login"], recipients[_itRec], message_text)
                tg_client.send_message(recipients[_itRec], message_text)
                bot.send_message(chat_id, "Отправлено сообщение юзеру " + recipients[
                    _itRec] + " текст - " + message_text + " от " + user["login"])
                file = open("clients/" + str(chat_id) + "/sended.txt", "a")
                file.write(recipients[_itRec] + "\n")
                file.close()
                _errors = 0
                sleep(2)
        except FileNotFoundError as e:
            file = open("clients/" + str(chat_id) + "/sended.txt", "w")
            file.close()
        except BaseException as e:
            bot.send_message(chat_id, "Не удалось отправить юзеру " + recipients[_itRec] + " от " + user[
                "login"] + " текст " + message_text + " " + str(e))
            _errors += 1
            if _errors != len(valid_users):
                _itRec -= 1
            else:
                _errors = 0
        tg_client.disconnect()
        _itRec += 1

@bot.message_handler(commands=['set_base'])
def set_base(message):
    chat_id = message.chat.id
    bot.send_message(chat_id, "Отправьте текстовый файл с логинами.")
    bot.register_next_step_handler(message, set_base_handler)

def set_base_handler(message):
    try:
        chat_id = message.chat.id

        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        if message.document.file_name.split(".")[1] != "xlsx":
            bot.send_message(chat_id, "Неверный формат файла, нужен xlsx")
            return
        src = "clients/"+str(chat_id)+"/base.xlsx";
        with open(src, 'wb') as new_file:
            new_file.write(downloaded_file)
    except Exception as e:
        bot.reply_to(message, e)

@bot.message_handler(commands=['delete_user'])
def delete_user(message):
    chat_id = message.chat.id
    bot.send_message(chat_id, "Введите номер аккаунта, который хотите удалить.")
    bot.register_next_step_handler(message, delete_user_handler)


def delete_user_handler(message):
    chat_id = message.chat.id
    login = message.text
    deleted = False
    for file_path in os.listdir("clients/" + str(chat_id) + "/users"):
        if file_path == login + ".clnt":
            os.remove("clients/" + str(chat_id) + "/users/" + file_path)
            deleted = True
    if deleted:
        bot.send_message(chat_id, "Удален.")
    else:
        bot.send_message(chat_id, "Не удалось найти аккаунт с логином " + login)

@bot.message_handler(commands=['clear_sended'])
def clear_sended(message):
    chat_id = message.chat.id
    file = open("clients/" + str(chat_id) + "/sended.txt", "w")
    file.close()
    bot.send_message(chat_id, "Очищено.")

def create_keyboard():
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(KeyboardButton('/set_message'))
    keyboard.add(KeyboardButton('/set_recipients'))
    keyboard.add(KeyboardButton('/send_spam'))
    keyboard.add(KeyboardButton("/user_append"))
    keyboard.add(KeyboardButton("/get_users"))
    keyboard.add(KeyboardButton("/delete_user"))
    keyboard.add(KeyboardButton("/set_base"))
    keyboard.add(KeyboardButton("/clear_sended"))
    keyboard.add(KeyboardButton("/stop_spam"))
    return keyboard


@bot.message_handler(func=lambda message: True)
def handle_message(message):
    chat_id = message.chat.id
    create_client_dir(chat_id)
    bot.send_message(chat_id, str(chat_id))
    bot.send_message(message.chat.id, 'Используйте кнопки ниже для управления.', reply_markup=create_keyboard())


def get_saved_users(client_id):
    users = []
    try:
        for file_path in os.listdir("clients/" + str(client_id) + "/users"):
            file = open("clients/" + str(client_id) + "/users/" + file_path, 'r')
            user_str = file.read()
            user = {}
            for line in user_str.split("\n"):
                if line.split("=")[0] == "api_id":
                    user["api_id"] = line.split("=")[1]
                if line.split("=")[0] == "api_hash":
                    user["api_hash"] = line.split("=")[1]
                if line.split("=")[0] == "login":
                    user["login"] = line.split("=")[1]
                if line.split("=")[0] == "password":
                    user["password"] = line.split("=")[1]
                if line.split("=")[0] == "proxy":
                    user["proxy"] = line.split("=")[1]
            users.append(user)
    except:
        bot.send_message(client_id, "У вас нет добавленных аккаунтов!")
    return users


if __name__ == '__main__':
    main()
